from celery import shared_task

from apps.core.models import (
    FileDetail,
    File,
    Report,
    Versioning
)

from django.conf import settings

import csv

import os

import pydash

from config.renderers import ExcelRenderer

import datetime

import xlsxwriter

import io

from xhtml2pdf import pisa

from django.template.loader import get_template

from openpyxl import load_workbook

from django.db.models import Count


def upload_excel(file):
    workbook = load_workbook(filename=f'{settings.MEDIA_ROOT}/{file.file_route}')

    sheet = workbook.active

    count = 0

    is_error = False

    position = None

    position_type_document = None

    position_social_reason = None

    position_surnames = None

    position_email = None

    position_department = None

    position_municipality = None

    position_country = None

    position_phone = None

    position_cell_phone = None

    position_address = None

    position_compound_address = None

    position_notification_address = None

    file_detail = []

    for row in sheet.iter_rows(values_only=True):
        if count == 0:
            if 'numero_documento' not in row:
                is_error = True
                break
            else:
                position = row.index('numero_documento')

            if 'tipo_documento' in row:
                position_type_document = row.index('tipo_documento')

            if 'nombre_razon_social' in row:
                position_social_reason = row.index('nombre_razon_social')

            if 'apellidos' in row:
                position_surnames = row.index('apellidos')

            if 'correo_electronico' in row:
                position_email = row.index('correo_electronico')

            if 'departamento' in row:
                position_department = row.index('departamento')

            if 'municipio' in row:
                position_municipality = row.index('municipio')

            if 'pais' in row:
                position_country = row.index('pais')

            if 'tel_fijo' in row:
                position_phone = row.index('tel_fijo')

            if 'tel_celular' in row:
                position_cell_phone = row.index('tel_celular')

            if 'direccion' in row:
                position_address = row.index('direccion')

            if 'direccion_notificacion' in row:
                position_notification_address = row.index('direccion_notificacion')
        else:
            try:
                FileDetail.objects.create(
                    file=file,
                    type_document=row[position_type_document] if position_type_document else 0,
                    number_document=row[position],
                    social_reason=row[position_social_reason] if position_social_reason else '',
                    surnames=row[position_surnames] if position_surnames else '',
                    email=row[position_email] if position_email else '',
                    department=row[position_department] if position_department else '',
                    municipality=row[position_municipality] if position_municipality else '',
                    country=row[position_country] if position_country else '',
                    phone=row[position_phone] if position_phone else '',
                    cell_phone=row[position_cell_phone] if position_cell_phone else '',
                    address=row[position_address] if position_address else '',
                    compound_address=row[position_compound_address] if position_compound_address else '',
                    notification_address=row[position_notification_address] if position_notification_address else ''
                )
            except Exception as error:
                print(error)

        count += 1


    if is_error:
        File.objects.filter(
            id=file.id
        ).update(
            status=File.FINISHED_ERRORS
        )
    else:
        File.objects.filter(
            id=file.id
        ).update(
            status=File.SUCCESSFULLY_COMPLETED
        )



# @shared_task
def upload_file():
    pendings = File.objects.filter(
        status=File.PENDING
    )

    for pending in pendings:
        _, extension = os.path.splitext(f'{settings.MEDIA_ROOT}/{pending.file_route}')

        if extension != '.csv':
            upload_excel(pending)
            return True

        try:
            with open(f'{settings.MEDIA_ROOT}/{pending.file_route}', newline='', encoding='utf-8') as csvfile:
                pending.status = File.IN_PROGRESS
                pending.save()

                reader = csv.reader(csvfile)

                count = 0

                position_type_document = None
                position_numero_documento = None
                position_nombre_razon_social = None
                position_apellidos = None
                position_dir_tipo = None
                position_dir_numero = None
                position_dir_apendice = None
                position_dir_orientacion = None
                position_dir_numero2 = None
                position_dir_apendice2 = None
                position_dir_orientacion2 = None
                position_dir_placa = None
                position_dir_interior = None
                position_dir_bloque = None
                position_direccion_especial = None
                position_correo_electronico = None
                position_departamento = None
                position_municipio = None
                position_pais = None
                position_tel_fijo = None
                position_tel_celular = None
                position_direccion = None
                position_direccion_notificacion = None
                file_detail = None
                is_error = False

                for row in reader:

                    new_data = [item for item in row]

                    if 'numero_documento' not in new_data:
                        new_data = [item.strip('"') for item in row[0].split(',')]

                    if count == 0:
                        if 'tipo_documento' in new_data:
                            position_type_document = new_data.index('tipo_documento')

                        if 'numero_documento' in new_data:
                            position_numero_documento = new_data.index('numero_documento')
                        else:
                            is_error = True
                            break

                        if 'nombre_razon_social' in new_data:
                            position_nombre_razon_social = new_data.index('nombre_razon_social')

                        if 'apellidos' in new_data:
                            position_apellidos = new_data.index('apellidos')

                        if 'dir_tipo' in new_data:
                            position_dir_tipo = new_data.index('dir_tipo')

                        if 'dir_numero' in new_data:
                            position_dir_numero = new_data.index('dir_numero')

                        if 'dir_apendice' in new_data:
                            position_dir_apendice = new_data.index('dir_apendice')

                        if 'dir_orientacion' in new_data:
                            position_dir_orientacion = new_data.index('dir_orientacion')

                        if 'dir_numero2' in new_data:
                            position_dir_numero2 = new_data.index('dir_numero2')

                        if 'dir_apendice2' in new_data:
                            position_dir_apendice2 = new_data.index('dir_apendice2')

                        if 'dir_orientacion2' in new_data:
                            position_dir_orientacion2 = new_data.index('dir_orientacion2')

                        if 'dir_placa' in new_data:
                            position_dir_placa = new_data.index('dir_placa')

                        if 'dir_interior' in new_data:
                            position_dir_interior = new_data.index('dir_interior')

                        if 'dir_bloque' in new_data:
                            position_dir_bloque = new_data.index('dir_bloque')

                        if 'direccion_especial' in new_data:
                            position_direccion_especial = new_data.index('direccion_especial')

                        if 'correo_electronico' in new_data:
                            position_correo_electronico = new_data.index('correo_electronico')

                        if 'departamento' in new_data:
                            position_departamento = new_data.index('departamento')

                        if 'municipio' in new_data:
                            position_municipio = new_data.index('municipio')

                        if 'pais' in new_data:
                            position_pais = new_data.index('pais')

                        if 'tel_fijo' in new_data:
                            position_tel_fijo = new_data.index('tel_fijo')

                        if 'tel_celular' in new_data:
                            position_tel_celular = new_data.index('tel_celular')

                        if 'direccion' in new_data:
                            position_direccion = new_data.index('direccion')

                        if 'direccion_notificacion' in new_data:
                            position_direccion_notificacion = new_data.index('direccion_notificacion')

                    elif count > 0:
                        address = build_address(
                            new_data[position_dir_tipo] if position_dir_tipo is not None else '',
                            new_data[position_dir_numero] if position_dir_numero is not None else '',
                            new_data[position_dir_apendice] if position_dir_apendice is not None else '',
                            new_data[position_dir_orientacion] if position_dir_orientacion is not None else '',
                            new_data[position_dir_numero2] if position_dir_numero2 is not None else '',
                            new_data[position_dir_apendice2] if position_dir_apendice2 is not None else '',
                            new_data[position_dir_orientacion2] if position_dir_orientacion2 is not None else '',
                            new_data[position_dir_placa] if position_dir_placa is not None else '',
                            new_data[position_dir_interior] if position_dir_interior is not None else '',
                            new_data[position_dir_bloque] if position_dir_bloque is not None else '',
                            new_data[position_direccion_especial] if position_direccion_especial is not None else ''
                        )

                        file_detail = [
                            FileDetail(
                                file=pending,
                                type_document=new_data[position_type_document] if position_type_document is not None else 0,
                                number_document=new_data[position_numero_documento] if position_numero_documento is not None else '',
                                social_reason=new_data[position_nombre_razon_social] if position_nombre_razon_social is not None else '',
                                surnames=new_data[position_apellidos] if position_apellidos is not None else '',
                                email=new_data[position_correo_electronico] if position_correo_electronico is not None else '',
                                department=new_data[position_departamento] if position_departamento is not None else '',
                                municipality=new_data[position_municipio] if position_municipio is not None else '',
                                country=new_data[position_pais] if position_pais is not None else '',
                                phone=new_data[position_tel_fijo] if position_tel_fijo is not None else '',
                                cell_phone=new_data[position_tel_celular] if position_tel_celular is not None else '',
                                address=new_data[position_direccion] if position_direccion is not None else '',
                                compound_address=address,
                                notification_address=new_data[position_direccion_notificacion] if position_direccion_notificacion is not None else ''
                            )
                        ]

                    count += 1
                    FileDetail.objects.bulk_create(file_detail)

            if is_error:
                pending.status = File.FINISHED_ERRORS
            else:
                pending.status = File.SUCCESSFULLY_COMPLETED

            pending.save()

        except Exception as error:
            print(error)
            pending.status = File.FINISHED_ERRORS
            pending.save()

def build_address(
    dir_tipo,
    dir_numero,
    dir_apendice,
    dir_orientacion,
    dir_numero2,
    dir_apendice2,
    dir_orientacion2,
    dir_placa,
    dir_interior,
    dir_bloque,
    direccion_especial
):
    return f'{dir_tipo} {dir_numero} {dir_apendice} {dir_orientacion} {dir_numero2} {dir_apendice2} {dir_orientacion2} {dir_placa} {dir_interior} {dir_bloque} {direccion_especial}'








@shared_task
def generate_excel_report(report_id):
    report = Report.objects.filter(id=report_id).first()

    workbook = xlsxwriter.Workbook(f'{settings.MEDIA_ROOT}/report_{report_id}.xlsx')
    worksheet = workbook.add_worksheet()

    worksheet.write(0, 1, 'NÚMERO DE DOCUMENTO')

    columns = list()

    if report.columns:
        from apps.core.serializers import FileDetailListGrouped

        columns = report.columns.split(',')

        if 'social_reason' in columns:
            worksheet.write(0, 2, 'RAZÓN SOCIAL')

        if 'address' in columns:
            worksheet.write(0, 4, 'DIRECCIÓN')

        if 'compound_address' in columns:
            worksheet.write(0, 5, 'DIRECICÓN COMPUESTA')

        if 'email' in columns:
            worksheet.write(0, 6, 'CORREO')

        if 'phone' in columns:
            worksheet.write(0, 10, 'TELÉFONO')

        if 'cell_phone' in columns:
            worksheet.write(0, 11, 'CELULAR')

        if 'notification_address' in columns:
            worksheet.write(0, 12, 'DIRECCIÓN DE NOTIFICACIÓN')

        worksheet.write(0, 13, 'CANTIDAD DE REGISTROS')

    else:
        worksheet.write(0, 2, 'RAZÓN SOCIAL')
        worksheet.write(0, 3, 'APELLIDOS')
        worksheet.write(0, 4, 'DIRECCIÓN')
        worksheet.write(0, 5, 'DIRECICÓN COMPUESTA')
        worksheet.write(0, 6, 'CORREO')
        worksheet.write(0, 7, 'DEPARTAMENTO')
        worksheet.write(0, 8, 'MUNICIPIO')
        worksheet.write(0, 9, 'PAIS')
        worksheet.write(0, 10, 'TELÉFONO')
        worksheet.write(0, 11, 'CELULAR')
        worksheet.write(0, 12, 'DIRECCIÓN DE NOTIFICACIÓN')

    if report is None:
        return False

    report.status = File.IN_PROGRESS
    report.file_excel = f'report_{report.id}.xlsx'
    report.save()

    details = FileDetail.objects.all().values()

    if report.document:
        documents = report.document.split(" ")

        details = details.filter(number_document__in=documents)

    if report.versioning:
        details = details.filter(file__versioning=report.versioning)

    columns.append('number_document')

    if report.columns:
        details = details.values(
            *columns
        ).annotate(
            count=Count('number_document')
        ).order_by(
            'number_document'
        )

    count = 1

    count_phone = 0

    count_cell_phone = 0

    count_address = 0

    count_email = 0

    for detail in details:

        phone = detail.get('phone', None)
        number_document = detail.get('number_document', None)
        social_reason = detail.get('social_reason', None)
        surnames = detail.get('surnames', None)
        address = detail.get('address', None)
        compound_address = detail.get('compound_address', None)
        email = detail.get('email', None)
        department = detail.get('department', None)
        municipality = detail.get('municipality', None)
        country = detail.get('country', None)
        cell_phone = detail.get('cell_phone', None)

        if phone:
            count_phone += 1

        if cell_phone:
            count_cell_phone += 1

        if email:
            count_email += 1

        if address:
            count_address += 1

        worksheet.write(count, 1, number_document)
        worksheet.write(count, 2, social_reason)
        worksheet.write(count, 3, surnames)
        worksheet.write(count, 4, address)
        worksheet.write(count, 5, compound_address)
        worksheet.write(count, 6, email)
        worksheet.write(count, 7, department)
        worksheet.write(count, 8, municipality)
        worksheet.write(count, 9, country)
        worksheet.write(count, 10, phone)
        worksheet.write(count, 11, cell_phone)
        worksheet.write(count, 12, compound_address)
        worksheet.write(count, 13, detail.get('count'))

        count += 1

    workbook.close()
    report.status = File.SUCCESSFULLY_COMPLETED
    report.number_records = count - 1

    report.phone = count_phone
    report.cell_phone = count_cell_phone
    report.address = count_address
    report.email = count_email

    report.save()

    generate_pdf_report(report)


def generate_pdf_report(report):
    template = get_template('report.html')

    files = None

    print('x,x,x,x,x,', report.files)
    print('x,x,x,x,x,', type(report.files))

    if report.files is not None:
        print('xxxx')
        print('xxxx')
        print('xxxx')
        print('xxxx')
        files = File.objects.filter(uuid__in=report.files.split(','))

    versioning = report.versioning
    versionings = Versioning.objects.all()

    print("files")
    print(files)

    html = template.render({
        'report': report,
        'files': files,
        'versioning': versioning,
        'all_versioning': versionings
    })


    route = os.path.join(settings.MEDIA_ROOT, 'reports/')
    os.makedirs(route, exist_ok=True)
    name_file = f'report_{report.id}.pdf'
    destination_path = os.path.join(route, name_file)

    destination = open(destination_path, 'wb')

    pisa.CreatePDF(
        html,
        dest=destination,
        link_callback=link_callback
    )

    report.file_pdf = f'reports/report_{report.id}.pdf'
    report.save()

    return destination_path


def link_callback(uri, rel):
    """
    Convert HTML URIs to absolute system paths so xhtml2pdf can access those
    resources
    """
    s_url = settings.STATIC_URL
    s_root = settings.STATIC_ROOT
    m_url = settings.MEDIA_URL
    m_root = settings.MEDIA_ROOT

    if uri.startswith(m_url):
        path = os.path.join(m_root, uri.replace(m_url, ""))
    elif uri.startswith(s_url):
        path = os.path.join(s_root, uri.replace(s_url, ""))
    else:
        return uri

    return path
